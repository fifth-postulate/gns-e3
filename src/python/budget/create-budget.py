from openpyxl import (Workbook, load_workbook)
from random import (shuffle, randint)

ITEM_LABEL_CELL = 'A1'
COST_LABEL_CELL = 'B1'
BUDGET_LABEL_CELL = 'C1'

RANGE_CELL = 'B2'
BUDGET_CELL = 'C2'

class Data:
    def __init__(self, minimum, maximum, header, subheader):
        self.minimum = minimum
        self.maximum = maximum
        self.header = header
        self.subheader = subheader
        self.data = []

    def add_item(self, item, price):
        self.data.append((item, price))

    def expand_to(self, ws):
        ws.title = 'BOM'
        self.header.expand_to(ws)
        self.subheader.expand_to(ws)
        for (item, price) in expand(self.minimum, self.maximum, self.data[:]):
            ws.append([item, price])

def expand(minimum, maximum, data):
    result = data + [(item, price) for (item, p) in data for price in [p + randint(100,10000) for _ in range(0, randint(minimum, maximum))]]
    shuffle(result)
    return result

class Header:
    def __init__(self, itemLabel, costLabel, budgetLabel):
        self.itemLabel = itemLabel
        self.costLabel = costLabel
        self.budgetLabel = budgetLabel

    def expand_to(self, ws):
        ws[ITEM_LABEL_CELL] = self.itemLabel
        ws[COST_LABEL_CELL] = self.costLabel
        ws[BUDGET_LABEL_CELL] = self.budgetLabel

class SubHeader:
    def __init__(self, range, budget):
        self.range = range
        self.budget = budget

    def expand_to(self, ws):
        ws[RANGE_CELL] = self.range
        ws[BUDGET_CELL] = self.budget

def read_data():
    base = load_workbook(filename='base.xlsx')
    bom = base['BOM']
    data = Data(
        50, 150,
        Header('Item', 'Cost', 'Budget'),
        SubHeader(bom[RANGE_CELL].value, bom[BUDGET_CELL].value)
    )
    for row in bom.iter_rows(min_row=3, max_row=10000, max_col=2):
        item = row[0].value
        price = row[1].value
        if not item:
            break
        data.add_item(item, price)
    return data

if __name__ == '__main__':
    data = read_data()
        
    wb = Workbook()
    ws = wb.active

    data.expand_to(ws)

    # Save the file
    wb.save('budget.xlsx')
